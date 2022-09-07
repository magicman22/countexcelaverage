import os
from openpyxl import Workbook, load_workbook
#cwd = os.getcwd()
#files = os.listdir(cwd)
#print("Files in %r: %s" % (cwd,files))

file_name = ["CGNP25_1P.xlsx", "CGNP25_1T.xlsx", "CGNP25_2P.xlsx", "CGNP25_2T.xlsx", "CGNP25_3P.xlsx", "CGNP25_3T.xlsx", "CGNP25_4P.xlsx", "CGNP25_4T.xlsx", "CGNP25_5P.xlsx", "CGNP25_5T.xlsx", "CGNP25_6P.xlsx", "CGNP25_6T.xlsx", "DW_1P.xlsx", "DW_1T.xlsx", "DW_2P.xlsx", "DW_2T.xlsx", "DW_3P.xlsx", "DW_3T.xlsx", "DW_4P.xlsx", "DW_4T.xlsx", "DW_5P.xlsx", "DW_5T.xlsx", "DW_6P.xlsx", "DW_6T.xlsx"]
for i in file_name:
    wb = load_workbook('Desktop/' + i)
    ws = wb.active
    ws['D7'] = '= AVERAGE(B56:B955)'
    averagevalue = ws['D7']
    wb.save('Desktop/' + i)
    #print(ws['D7'].value) -> dont have to print for now


